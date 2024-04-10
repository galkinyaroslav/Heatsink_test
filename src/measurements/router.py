import asyncio
import json
import os.path
from datetime import datetime, UTC
from fastapi import APIRouter, Depends, Response, BackgroundTasks, WebSocket, WebSocketDisconnect
from fastapi.websockets import WebSocketState
from pydantic import BaseModel

from sqlalchemy import select, insert, update, func, Row, RowMapping
from sqlalchemy.ext.asyncio import AsyncSession

from auth.auth_config import fastapi_users
from auth.models import User
from database import get_async_session
import measurements.meas_control as mc
from measurements.models import Run, Measurement
from measurements.schemas import RunPart, MeasurementsToDB
import csv
from openpyxl import Workbook
from measurements.models import Run
from measurements.plot_gen import gen_hex_plot

a34970 = mc.find_device()

current_user = fastapi_users.current_user()

router = APIRouter(
    prefix='/measurements',
    tags=['Measurement']
)


@router.post('/new-run', name='new_run')
async def new_run(session: AsyncSession = Depends(get_async_session),
                  user: User = Depends(current_user)):
    # update Run.number incrementing by 1
    stmt = update(Run).filter(Run.id == 1).values(number=Run.number + 1)
    await session.execute(stmt)
    await session.commit()
    query = select(Run).where(Run.id == 1)
    result = await session.scalars(query)
    run = result.one()
    return run.number


@router.post('/get-run-num', name='get_run_number')
async def get_run_num(session: AsyncSession = Depends(get_async_session),
                      user: User = Depends(current_user)):
    # get current Run.number
    query = select(Run).where(Run.id == 1)
    result = await session.scalars(query)
    run = result.one()
    if result is None:
        first_run = Run(id=1, number=1)
        session.add(first_run)
        # print('add one')
        await session.commit()
        return first_run.number
    else:
        return run.number


@router.get('/configure')
async def configurate_device(session: AsyncSession = Depends(get_async_session),
                             user: User = Depends(current_user)):
    # configure device
    data = mc.configure(a34970)
    return data


@router.get('/get-measurements')
async def get_measurements(session: AsyncSession = Depends(get_async_session),
                           user: User = Depends(current_user)):
    # data from device
    data = mc.read_data(a34970)
    return data


@router.post('/run-meas', name='run_meas')
async def run_meas(part: RunPart = RunPart.first20,
                   session: AsyncSession = Depends(get_async_session),
                   user: User = Depends(current_user)):
    run = await session.scalars(select(Run).where(Run.id == 1))
    data = mc.read_data(a34970)  # TODO UNCOMMENT IT
    # data = {'data': [float(i) for i in range(20)]}  # TODO COMMENT IT!!!!
    measurements_to_db = Measurement()
    measurements_to_db.data = data['data']
    measurements_to_db.run_number = run.one().number
    measurements_to_db.user_id = user.id
    measurements_to_db.measuredpart = part.name
    session.add(measurements_to_db)
    await session.commit()
    await session.refresh(measurements_to_db)
    return measurements_to_db


@router.get('/get-runs')
async def get_runs(offset: int | None = None,
                   limit: int | None = None,
                   session: AsyncSession = Depends(get_async_session),
                   user: User = Depends(current_user)):
    query = (select(Measurement.run_number, func.min(Measurement.measure_datetime).label('datetime'))
             .group_by(Measurement.run_number)
             .offset(offset)
             .limit(limit)
             .order_by(Measurement.run_number)
             )
    results = await session.execute(query)
    run_numbers = results.mappings()
    # print(run_numbers)
    response = [i for i in run_numbers]
    # print(response)
    return {'lostofdata': response,
            'user': user}


@router.get('/runs/{run_number}')
async def get_arun(run_number: int,
                   session: AsyncSession = Depends(get_async_session),
                   user: User = Depends(current_user)):
    # data from db
    query = select(Measurement).where(Measurement.run_number == run_number)
    results = await session.execute(query)
    listofdata = results.scalars().all()
    return {'listofdata': listofdata,
            'user': user}


@router.post('/runs/{run_number}/data-to-csv', name='arun_to_csv')
async def data_to_csv(run_number: int,
                      session: AsyncSession = Depends(get_async_session),
                      user: User = Depends(current_user)):
    # data from db
    query = select(Measurement).where(Measurement.run_number == run_number)
    results = await session.execute(query)
    listofdata = results.scalars().all()
    path = './saved'
    print(os.getcwd())
    if not os.path.exists(path):
        os.makedirs(path)

    outfile = open(f'{path}/#{run_number}.csv', 'w')
    outcsv = csv.writer(outfile)
    header = ['part'] + [f'ch#{i}' for i in range(1, 21, 1)]
    header.append('datetime')
    outcsv.writerow(header)

    for i in listofdata:
        data = [i.measuredpart] + i.data
        data.append(i.measure_datetime)
        print()
        outcsv.writerow(data)

    outfile.close()
    return {'message': 'Saved'}


@router.post('/runs/{run_number}/data-to-excel', name='arun_to_excel')
async def data_to_excel(run_number: int,
                        session: AsyncSession = Depends(get_async_session),
                        user: User = Depends(current_user)):
    # data from db
    query1 = select(Measurement).where(Measurement.run_number == run_number,
                                       Measurement.measuredpart == RunPart.first20)
    query2 = select(Measurement).where(Measurement.run_number == run_number,
                                       Measurement.measuredpart == RunPart.second20)
    query3 = select(Measurement).where(Measurement.run_number == run_number,
                                       Measurement.measuredpart == RunPart.third20)
    results1 = await session.execute(query1)
    listofdata1 = results1.scalars().all()
    results2 = await session.execute(query2)
    listofdata2 = results2.scalars().all()
    results3 = await session.execute(query3)
    listofdata3 = results3.scalars().all()
    header_bottom = ['timestamp'] + [i for i in range(1, 24, 1)]
    header_top = ['timestamp'] + [i for i in range(24, 59, 1)]
    header_water = ['timestamp'] + [59, 60]
    wb = Workbook()
    ws_bottom = wb.active
    ws_bottom.title = 'Bottom side'
    ws_top = wb.create_sheet('Top side')
    ws_water = wb.create_sheet('Water')
    ws_bottom.append(header_bottom)
    ws_top.append(header_top)
    ws_water.append(header_water)
    delta_bottom = 0
    delta_top = 0
    delta_watter = 0
    for i, j, k in zip(listofdata1, listofdata2, listofdata3):
        # print(i.data[18], i.data[19])
        if delta_bottom == 0:
            time_bottom = 0
            time_top = 0
            time_water = 0
            delta_bottom = i.measure_datetime.timestamp()
            delta_top = j.measure_datetime.timestamp()
            delta_watter = k.measure_datetime.timestamp()
        else:
            time_bottom = i.measure_datetime.timestamp() - delta_bottom
            time_top = j.measure_datetime.timestamp() - delta_top
            time_water = k.measure_datetime.timestamp() - delta_watter
        row_bottom = [time_bottom] + i.data[:] + j.data[:3]
        ws_bottom.append(row_bottom)
        row_top = [time_top] + j.data[3:] + k.data[:18]
        ws_top.append(row_top)
        row_water = [time_water] + k.data[18:]
        ws_water.append(row_water)

    wb.save(f'saved/#{run_number}.xlsx')
    wb.close()
    return {'message': 'Saved'}


@router.post('/run-meas-top')
async def run_meas_top(session: AsyncSession = Depends(get_async_session)):
    # run measurements top side and add it to db
    return


@router.get('/get-hex-plot')
async def get_hex_plot(side: str = 'top',
                       session: AsyncSession = Depends(get_async_session),
                       user: User = Depends(current_user)):
    # run measurements top side and add it to db
    data = []
    buf = gen_hex_plot(data=data, side=side)
    response = Response(content=buf.getvalue(), media_type='image/png')
    return response

# WEBSOCKETS
class ConnectionManager:
    def __init__(self):
        self.active_connections: set[WebSocket] = set()

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.add(websocket)

    def disconnect(self, websocket: WebSocket):
        # websocket.close()
        self.active_connections.remove(websocket)

    async def send_dataset(self, message: json, websocket: WebSocket):
        await websocket.send_json(message)

    async def broadcast(self, message: json):
        for connection in self.active_connections:
            await connection.send_json(message)


stop_timer = False
data_ready = False
measured_data = []
measurement_time = []

# Функция для обработки WebSocket-соединения
# @router.websocket('/ws')
# async def websocket_endpoint(websocket: WebSocket):
#     await websocket.accept()
#     global stop_timer
#     while not stop_timer:
#         if websocket.client_state == WebSocketState.DISCONNECTED:
#             print(websocket.client_state)
#             break
#         # print(chart_data)
#         await websocket.send_json(chart_data[-1])
#         await asyncio.sleep(5)
#     await websocket.close()
#     stop_timer = False

manager = ConnectionManager()


class WebSocketData:
    def __init__(self, temperature: list[str] = None):
        self.temperature = temperature


ws_data = WebSocketData()


@router.websocket('/ws')
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    global stop_timer
    global data_ready
    print(websocket.client_state)
    try:
        while True:
            if data_ready:
                # data = await websocket.receive_text()
                # await manager.send_dataset(ws_data.temperature, websocket)
                await manager.broadcast(message={'data': measured_data[-1],
                                                 'time': measurement_time[-1]})
                data_ready = False

            await asyncio.sleep(1)
            # await manager.broadcast(f"Client # says: {data}")
    except Exception as e:
        print(f'Error occurred {e}')
        # await manager.broadcast(f"Client #{websocket} left the chat")
    finally:
        manager.disconnect(websocket)
        print(f'{websocket} is disconnected')


@router.post("/start-timer1/")
async def start_timer_task(background_tasks: BackgroundTasks):
    global stop_timer
    stop_timer = False  # Reset the stop flag

    global measurement_time
    measurement_time = [0,]
    global measured_data
    measured_data = []
    previous_time = 0
    global data_ready

    wb = Workbook()
    ws = wb.active
    ws.title = 'RT'

    path = './saved'
    # print(os.getcwd())
    if not os.path.exists(path):
        os.makedirs(path)
    wb.save(f'saved/RT{str(datetime.now(UTC))}.xlsx')
    wb.close()
    while not stop_timer:
        data_ready = False
        if previous_time == 0:
            # current_time = 0
            previous_time = int(datetime.timestamp(datetime.utcnow()))
        else:
            current_time = int(datetime.timestamp(datetime.utcnow()) - previous_time)
            measurement_time.append(current_time)
        meas_data = mc.read_data(a34970)['data']
        measured_data.append(meas_data)

        data_ready = True
        print(meas_data)
        await asyncio.sleep(5)


@router.post("/stop-timer1/")
async def stop_timer_task():
    global measurement_time
    global measured_data
    print(measured_data)
    print(measurement_time)
    # wb = Workbook()
    # ws = wb.active
    # ws.title = 'RT'
    # row = [time_bottom] + i.data[:] + j.data[:3]
    #     for i in range(len(measurement_time)):
    #         row = [measurement_time[i]]+measured_data[i]
    #         print(row)
    #         ws.append(row)
    # path = './saved'
    # # print(os.getcwd())
    # if not os.path.exists(path):
    #     os.makedirs(path)
    # wb.save(f'saved/RT{str(datetime.utcnow())}.xlsx')
    # wb.close()
    measurement_time = [0,]
    measured_data = []
    global stop_timer
    stop_timer = True
    global data_ready
    data_ready = False
    return {"message": "Timer stopped"}
